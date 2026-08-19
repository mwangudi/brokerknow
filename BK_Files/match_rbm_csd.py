"""Match the RBM client-accounts report to our clients and emit the apply SQL.

Tiering by strength of evidence, because a wrong CSD number routes someone
else's trades to the wrong client:

  A  identity document matches exactly                    -> apply
  B  name matches exactly AND the documents are near-identical
     (one contains the other, or differs by <= 2 chars)    -> apply
  C  name matches but the documents clearly disagree       -> REPORT ONLY
  D  same name held by more than one client                -> REPORT ONLY
  E  no counterpart in our book                            -> REPORT ONLY

Writes: BK_Files/apply_cedar_csd_0818.sql   (gitignored - contains client PII)
        tmp/csd_review_0818.csv             (tiers C/D for a human to settle)
"""
import re, csv, unicodedata
from collections import defaultdict
import openpyxl

XLSX = r"c:\Users\v-mwangudi\source\repos\BrokerKnow\BK_Files\CEDAR CAPITAL CLIENT LIST 18-08-2026.xlsx"
PSV = r"c:\Users\v-mwangudi\source\repos\BrokerKnow\tmp\clients_0818.psv"
SQL_OUT = r"c:\Users\v-mwangudi\source\repos\BrokerKnow\BK_Files\apply_cedar_csd_0818.sql"
CSV_OUT = r"c:\Users\v-mwangudi\source\repos\BrokerKnow\tmp\csd_review_0818.csv"


def norm_name(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).upper()
    s = re.sub(r"\b(MR|MRS|MS|MISS|DR|PROF|REV|SIR)\b", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_id(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper()) if s else ""


def uid_to_id(uid):
    u = norm_id(uid)
    return u[2:] if u.startswith("PP") else u


def close(a, b):
    """Same document recorded slightly differently."""
    if not a or not b or len(a) < 5 or len(b) < 5:
        return False
    if a in b or b in a:
        return True
    if abs(len(a) - len(b)) > 1:
        return False
    pad = max(len(a), len(b))
    a2, b2 = a.ljust(pad), b.ljust(pad)
    return sum(1 for x, y in zip(a2, b2) if x != y) <= 2


clients = []
for line in open(PSV, encoding="utf-8", errors="replace"):
    p = line.rstrip("\n").split("|")
    if len(p) < 4 or not p[0].strip().isdigit():
        continue
    clients.append({"dpa": int(p[0].strip()), "name": p[1].strip(),
                    "idp": p[2].strip(), "csd": p[3].strip()})

by_id, by_name = defaultdict(list), defaultdict(list)
for c in clients:
    if norm_id(c["idp"]):
        by_id[norm_id(c["idp"])].append(c)
    if norm_name(c["name"]):
        by_name[norm_name(c["name"])].append(c)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rbm = []
for r in ws.iter_rows(min_row=6, values_only=True):
    code = str(r[0]).strip() if r[0] else ""
    if not code.startswith("CEDA"):
        continue
    rbm.append({"code": code,
                "type": str(r[2]).strip() if r[2] else "",
                "uid": str(r[3]).strip() if r[3] else "",
                "name": str(r[5]).strip() if r[5] else "",
                "default": bool(r[14]) if len(r) > 14 else False,
                "status": str(r[15]).strip() if len(r) > 15 and r[15] else ""})

tierA, tierB, tierC, tierD, tierE = [], [], [], [], []
for a in rbm:
    aid = uid_to_id(a["uid"])
    cand = by_id.get(aid, [])
    if len(cand) == 1:
        tierA.append((a, cand[0])); continue
    if len(cand) > 1:
        tierD.append((a, cand)); continue
    cand = by_name.get(norm_name(a["name"]), [])
    if not cand:
        tierE.append(a)
    elif len(cand) > 1:
        tierD.append((a, cand))
    elif close(aid, norm_id(cand[0]["idp"])):
        tierB.append((a, cand[0]))
    else:
        tierC.append((a, cand[0]))

# A client with several CSD accounts: prefer the Active default one.
chosen = {}
conflict = defaultdict(list)
for a, c in tierA + tierB:
    conflict[c["dpa"]].append(a)
for dpa, accts in conflict.items():
    if len(accts) == 1:
        chosen[dpa] = accts[0]
    else:
        pref = [x for x in accts if x["default"] and x["status"].lower() == "active"] \
               or [x for x in accts if x["status"].lower() == "active"] or accts
        chosen[dpa] = sorted(pref, key=lambda x: x["code"])[0]

multi = {d: a for d, a in conflict.items() if len(a) > 1}

print(f"RBM accounts            : {len(rbm)}")
print(f"our clients             : {len(clients)}")
print()
print(f"A  document match       : {len(tierA)}")
print(f"B  name + near document : {len(tierB)}")
print(f"C  name, document differs: {len(tierC)}   -> review, NOT applied")
print(f"D  ambiguous name/doc   : {len(tierD)}   -> review, NOT applied")
print(f"E  not in our book      : {len(tierE)}   -> informational")
print()
print(f"clients that will receive a CSD number : {len(chosen)}")
print(f"clients holding several CSD accounts   : {len(multi)} (Active default chosen)")

with open(SQL_OUT, "w", encoding="utf-8") as f:
    f.write("/* CSD numbers from 'CEDAR CAPITAL CLIENT LIST 18-08-2026.xlsx' (RBM client\n"
            "   accounts report). That file carries no client code, so clients were matched\n"
            "   on identity document, or on an exact name with a near-identical document.\n"
            "   Name-only matches whose documents disagree are deliberately EXCLUDED. */\n"
            "SET NOCOUNT ON; SET XACT_ABORT ON;\n"
            "IF COL_LENGTH('dbo.Client','ClientCDSNo') IS NOT NULL\n"
            "  AND (SELECT CHARACTER_MAXIMUM_LENGTH FROM INFORMATION_SCHEMA.COLUMNS\n"
            "       WHERE TABLE_NAME='Client' AND COLUMN_NAME='ClientCDSNo') < 50\n"
            "  ALTER TABLE dbo.Client ALTER COLUMN ClientCDSNo nvarchar(50) NULL;\nGO\n"
            "IF OBJECT_ID('dbo.ClientCDS_backup_0818') IS NULL\n"
            "  SELECT Client_DPA_, ClientCDSNo INTO dbo.ClientCDS_backup_0818 FROM dbo.Client;\nGO\n"
            "CREATE TABLE #csd (dpa int PRIMARY KEY, code nvarchar(50));\n")
    for dpa in sorted(chosen):
        f.write(f"INSERT INTO #csd VALUES ({dpa}, N'{chosen[dpa]['code']}');\n")
    f.write("\nBEGIN TRAN;\n"
            "UPDATE c SET c.ClientCDSNo = s.code\n"
            "FROM dbo.Client c JOIN #csd s ON s.dpa = c.Client_DPA_\n"
            "WHERE ISNULL(c.Deleted,0)=0;\n"
            "DECLARE @n int = @@ROWCOUNT;\n"
            f"PRINT 'expected : {len(chosen)}';\n"
            "PRINT 'updated  : ' + CAST(@n AS varchar(10));\n"
            "SELECT COUNT(*) AS truncated FROM dbo.Client WHERE ClientCDSNo IS NOT NULL AND LEN(ClientCDSNo)<>23;\n"
            "COMMIT;\nGO\n"
            "SELECT COUNT(*) AS clients, SUM(CASE WHEN ISNULL(ClientCDSNo,'')<>'' THEN 1 ELSE 0 END) AS with_csd\n"
            "FROM dbo.Client WHERE ISNULL(Deleted,0)=0;\nGO\n")

with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["tier", "reason", "csd_code", "rbm_name", "rbm_uid",
                "our_client", "our_name", "our_id"])
    for a, c in tierC:
        w.writerow(["C", "name matches, document differs", a["code"], a["name"],
                    a["uid"], f"C-{c['dpa']}", c["name"], c["idp"]])
    for a, cs in tierD:
        w.writerow(["D", f"matches {len(cs)} clients", a["code"], a["name"], a["uid"],
                    " / ".join(f"C-{x['dpa']}" for x in cs),
                    " / ".join(x["name"] for x in cs), ""])
    for dpa, accts in multi.items():
        nm = next(c["name"] for c in clients if c["dpa"] == dpa)
        w.writerow(["M", f"client has {len(accts)} CSD accounts; used {chosen[dpa]['code']}",
                    " / ".join(x["code"] for x in accts), "", "", f"C-{dpa}", nm, ""])

print(f"\nwrote {SQL_OUT}")
print(f"wrote {CSV_OUT}  ({len(tierC)+len(tierD)+len(multi)} rows to review)")
