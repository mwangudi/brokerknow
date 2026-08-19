"""CSV of clients still without a CDS number, for the team to complete.

Pre-fills a suggested number wherever the RBM register has a name match we
withheld (tier C/D), so the team verifies rather than researches from scratch.
Deliberately does NOT guess on weak evidence - a blank is safer than a wrong
CDS number, which would route someone else's trades to this client.

Reads : tmp/nocsd.psv (live clients with no CDS), the RBM xlsx
Writes: tmp/CDS_numbers_to_complete.csv   (client PII - tmp/ is gitignored)
"""
import re, csv, unicodedata
from collections import defaultdict
import openpyxl

BASE = r"c:\Users\v-mwangudi\source\repos\BrokerKnow"
XLSX = BASE + r"\BK_Files\CEDAR CAPITAL CLIENT LIST 18-08-2026.xlsx"
PSV = BASE + r"\tmp\nocsd.psv"
OUT = BASE + r"\tmp\CDS_numbers_to_complete.csv"


def norm_name(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).upper()
    s = re.sub(r"\b(MR|MRS|MS|MISS|DR|PROF|REV|SIR)\b", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_id(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper()) if s else ""


def uid_to_id(uid):
    u = norm_id(uid)
    return u[2:] if u.startswith("PP") else u


wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rbm = []
for r in ws.iter_rows(min_row=6, values_only=True):
    code = str(r[0]).strip() if r[0] else ""
    if not code.startswith("CEDA"):
        continue
    rbm.append({"code": code,
                "uid": str(r[3]).strip() if r[3] else "",
                "name": str(r[5]).strip() if r[5] else "",
                "default": bool(r[14]) if len(r) > 14 else False,
                "status": str(r[15]).strip() if len(r) > 15 and r[15] else ""})

by_name, by_tokens = defaultdict(list), defaultdict(list)
for a in rbm:
    n = norm_name(a["name"])
    if n:
        by_name[n].append(a)
        by_tokens[frozenset(n.split())].append(a)

rows = []
for line in open(PSV, encoding="utf-8", errors="replace"):
    p = line.rstrip("\n").split("|")
    if len(p) < 11 or not p[0].strip().isdigit():
        continue
    dpa, name, idp, doctype, acctype, reg, email, cell, trades, last, cash = \
        [x.strip() for x in p[:11]]

    n = norm_name(name)
    cand = by_name.get(n, [])
    how = "name on the register matches exactly"
    if not cand:
        cand = by_tokens.get(frozenset(n.split()), []) if n else []
        how = "same name parts on the register, in a different order"

    sugg = rbm_name = rbm_id = rbm_status = ""
    if len(cand) == 1:
        a = cand[0]
        sugg, rbm_name, rbm_id = a["code"], a["name"], uid_to_id(a["uid"])
        why = (f"NOT APPLIED: {how}, but the ID/passport differs "
               f"(ours '{idp}', register '{rbm_id}'). Please confirm this is the same person.")
        rbm_status = a["status"]
    elif len(cand) > 1:
        why = (f"NOT APPLIED: {len(cand)} accounts on the register carry this name "
               f"({', '.join(x['code'] for x in cand[:4])}). Please tell us which one.")
    else:
        why = "No account on the register matched this client by ID/passport or by name."

    ntrades, ncash = int(trades or 0), float(cash or 0)
    if re.search(r"\((BUY|SELL)\)$", name.strip(), re.I):
        prio = "0 - house/counterparty account, may not need a CDS number"
    elif ntrades and last >= "2025-08":
        prio = "1 - traded in the last year"
    elif ntrades:
        prio = "2 - has traded before"
    elif ncash:
        prio = "3 - holds a cash balance"
    else:
        prio = "4 - dormant, never traded, no balance"

    rows.append({
        "Priority": prio,
        "Client Code": dpa, "Client Name": name, "ID/Passport (our record)": idp,
        "ID Doc Type": doctype, "Account Type": acctype, "Registered": reg,
        "Email": email, "Mobile": cell, "Trades": int(trades or 0),
        "Last Trade": last, "Cash Balance": cash,
        "CDS NUMBER (please complete)": "",
        "Suggested CDS Number": sugg, "Name on register": rbm_name,
        "ID on register": rbm_id, "Register status": rbm_status,
        "Why it was not applied": why,
    })

rows.sort(key=lambda r: (r["Priority"], -r["Trades"], r["Client Name"]))

cols = list(rows[0].keys())
with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=cols)
    w.writeheader()
    w.writerows(rows)

withsugg = sum(1 for r in rows if r["Suggested CDS Number"])
multi = sum(1 for r in rows if "Please tell us which one" in r["Why it was not applied"])
print(f"clients without a CDS number : {len(rows)}")
print(f"  with a suggested number     : {withsugg}")
print(f"  register has several matches: {multi}")
print(f"  no candidate at all         : {len(rows) - withsugg - multi}")
for p in sorted({r["Priority"] for r in rows}):
    print(f"  {p:<52} {sum(1 for r in rows if r['Priority'] == p)}")
print(f"\nwrote {OUT}")
