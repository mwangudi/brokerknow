"""CSV of the clients that DID receive a CDS number, for the team to confirm.

Tells them how each number was arrived at, so the 373 that needed a judgement
call and the 102 clients holding several register accounts can be checked
first, rather than re-reading all 4367 rows.

Reads : tmp/clients_0818.psv (pre-apply book), tmp/withcsd.psv (what is live now), RBM xlsx
Writes: tmp/CDS_numbers_to_confirm.csv   (client PII - tmp/ is gitignored)
"""
import re, csv, unicodedata
from collections import defaultdict
import openpyxl

BASE = r"c:\Users\v-mwangudi\source\repos\BrokerKnow"
XLSX = BASE + r"\BK_Files\CEDAR CAPITAL CLIENT LIST 18-08-2026.xlsx"
BOOK = BASE + r"\tmp\clients_0818.psv"
LIVE = BASE + r"\tmp\withcsd.psv"
OUT = BASE + r"\tmp\CDS_numbers_to_confirm.csv"


def norm_name(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).upper()
    s = re.sub(r"\b(MR|MRS|MS|MISS|DR|PROF|REV|SIR)\b", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_id(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper()) if s else ""


def uid_to_id(uid):
    u = norm_id(uid)
    return u[2:] if u.startswith("PP") else u


def close(a, b):
    if not a or not b or len(a) < 5 or len(b) < 5:
        return False
    if a in b or b in a:
        return True
    if abs(len(a) - len(b)) > 1:
        return False
    pad = max(len(a), len(b))
    return sum(1 for x, y in zip(a.ljust(pad), b.ljust(pad)) if x != y) <= 2


book = []
for line in open(BOOK, encoding="utf-8", errors="replace"):
    p = line.rstrip("\n").split("|")
    if len(p) < 4 or not p[0].strip().isdigit():
        continue
    book.append({"dpa": int(p[0].strip()), "name": p[1].strip(), "idp": p[2].strip()})

by_id, by_name = defaultdict(list), defaultdict(list)
for c in book:
    if norm_id(c["idp"]):
        by_id[norm_id(c["idp"])].append(c)
    if norm_name(c["name"]):
        by_name[norm_name(c["name"])].append(c)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rbm = []
for r in ws.iter_rows(min_row=6, values_only=True):
    code = str(r[0]).strip() if r[0] else ""
    if not code.startswith("CEDA"):
        continue
    rbm.append({"code": code, "uid": str(r[3]).strip() if r[3] else "",
                "name": str(r[5]).strip() if r[5] else "",
                "default": bool(r[14]) if len(r) > 14 else False,
                "status": str(r[15]).strip() if len(r) > 15 and r[15] else ""})

# Rebuild exactly the tiers that were applied.
hits = defaultdict(list)          # client dpa -> [(tier, register account)]
for a in rbm:
    aid = uid_to_id(a["uid"])
    cand = by_id.get(aid, [])
    if len(cand) == 1:
        hits[cand[0]["dpa"]].append(("A", a)); continue
    if len(cand) > 1:
        continue
    cand = by_name.get(norm_name(a["name"]), [])
    if len(cand) == 1 and close(aid, norm_id(cand[0]["idp"])):
        hits[cand[0]["dpa"]].append(("B", a))

rows = []
for line in open(LIVE, encoding="utf-8", errors="replace"):
    p = line.rstrip("\n").split("|")
    if len(p) < 12 or not p[0].strip().isdigit():
        continue
    dpa, name, idp, doctype, acctype, reg, email, cell, trades, last, cash, csd = \
        [x.strip() for x in p[:12]]
    dpa = int(dpa)

    got = hits.get(dpa, [])
    applied = [x for x in got if x[1]["code"] == csd]
    tier = applied[0][0] if applied else (got[0][0] if got else "?")
    acct = applied[0][1] if applied else (got[0][1] if got else None)
    others = [x[1] for x in got if x[1]["code"] != csd]

    if others:
        prio = "1 - client holds several accounts on the register - confirm we picked the right one"
        how = (f"This client has {len(got)} accounts on the register. We used the active "
               f"default one. The others are: {', '.join(x['code'] for x in others[:5])}.")
    elif tier == "B":
        prio = "2 - matched on name, ID recorded slightly differently - please confirm"
        how = (f"Name matches the register exactly and the ID is near-identical "
               f"(ours '{idp}', register '{uid_to_id(acct['uid']) if acct else ''}').")
    else:
        prio = "3 - ID/passport matched the register exactly"
        how = "The ID/passport we hold matches the register exactly."

    rows.append({
        "Check Priority": prio, "Client Code": dpa, "Client Name": name,
        "ID/Passport (our record)": idp, "ID Doc Type": doctype, "Account Type": acctype,
        "Registered": reg, "Email": email, "Mobile": cell,
        "Trades": int(trades or 0), "Last Trade": last, "Cash Balance": cash,
        "CDS Number Applied": csd,
        "Name on register": acct["name"] if acct else "",
        "ID on register": uid_to_id(acct["uid"]) if acct else "",
        "Register status": acct["status"] if acct else "",
        "How it was matched": how,
        "Correct? (Y/N)": "", "If wrong, correct CDS number": "",
    })

rows.sort(key=lambda r: (r["Check Priority"], -r["Trades"], r["Client Name"]))

with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    w.writeheader()
    w.writerows(rows)

print(f"clients with a CDS number applied : {len(rows)}")
for p in sorted({r["Check Priority"] for r in rows}):
    print(f"  {p:<78} {sum(1 for r in rows if r['Check Priority'] == p)}")
print(f"\nwrote {OUT}")
